import os
import json
from openai import AzureOpenAI
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

CRITERIA = [
  {"S.No.": "1", "Particular": "Proof of Tender fees and EMD paid "},
  {"S.No.": "2", "Particular": "Copy of EMD exemption certificate if applicable. "},
  {"S.No.": "3", "Particular": "Attested photocopy of valid manufacturing drug license with product list duly approved by the central licensing authority/ state licencing authority for each and every product quoted as per technical specification in the bid. In case of importers, a valid licence of import and sale (in Form 10 with Form 41) shall be submitted. The license must have been duly renewed up to date and the quoted items along with drug code in the tender shall be clearly highlighted in the license. If quoted item is manufactured at different places, Manufacturing License & Performance certificate from all such places from respective Licensing Authority/State Drug Authority should be enclosed. For the purpose of this tender, it is hereby clarified that a Bidder could be a Loan Licensee for manufacturing drugs. Where the bidder participates as a Loan Licensee, submission of a valid certificate issued by the Principal Manufacturer shall be mandatory."},
  {"S.No.": "4", "Particular": "Valid World Health Organization-Good Manufacturing Practice (WHO-GMP) certificate or COPP highlighting the quoted products for importer."},
  {"S.No.": "5", "Particular": "Copy of permission from DCGI for 'New drug & Fixed Dose Combination'."},
  {"S.No.": "6", "Particular": "Market Standing Certificate issued by the Licensing Authority as a Manufacturer for each drug quoted for the last 3 year (i.e., for financial year 2022-23 and 2023-24 and 2024-25 and Certificate obtained specifically for CGMSC tender / it should be in General) should be uploaded with list of drugs. In case of direct importer, evidence of import of the said drugs for the last 3 years such as bill of landing, bill of entry for last 3 years and certificate of analysis are to be produced (irrespective of the Importer). In case the market standing certificate is not having the drugs as per the specifications mentioned in the tender, it will not be considered for further processing.  In cases involving new drugs/ drugs out of patent period it is sufficient to possess relevant market standing certificate, as applicable. Market Standing Certificate (MSC) of Manufacturing unit of Loan Licensee for 3 years is needed. \nNote: Document should be uploaded in colour scan copy of either original or Notarized copy."},
  {"S.No.": "7", "Particular": "Appendix-B (Checklist)"},
  {"S.No.": "8", "Particular": "Annexure-1 (Technical Specifications and Compliance)"},
  {"S.No.": "9", "Particular": "Annexure-2 (Letter Comprising Technical Bid)"},
  {"S.No.": "10", "Particular": "Annexure-3 (Proforma for Production and Sale Statement) "},
  {"S.No.": "11", "Particular": "Annexure-4 (Details of Manufacturing Unit)"},
  {"S.No.": "12", "Particular": "Annexure-5 (Details of Items Quoted with Drug Code)"},
  {"S.No.": "13", "Particular": "Annexure-6 (Annual Turnover Statement for drug related business for three (3) Years) along with Copies of Balance Sheet and Profit and Loss Accounts for last three years i.e., (2022-23, 2023-24, 2024-25) certified by the Statutory Auditor or Chartered Accountant. "},
  {"S.No.": "14", "Particular": "GST Registration certificate along with copy of the GST return not older than 3 months from the bid opening date."},
  {"S.No.": "15", "Particular": "Annexure-7 (Format of Power of Attorney for signing of Bid) except for proprietorship."},
  {"S.No.": "16", "Particular": "Annexure-8 (Affidavit for Blacklisting)"},
  {"S.No.": "17", "Particular": "Annexure-9 (Non-Conviction certificate)"},
  {"S.No.": "18", "Particular": "Annexure-10 (Mandate Form)"},
  {"S.No.": "19", "Particular": "Incorporation / Registration Certificate of Bidder"},
  {"S.No.": "20", "Particular": "Annexure- 12 (Bank Guarantee format for EMD), if applicable"},
  {"S.No.": "21", "Particular": "Annexure-13 (Pre-Contract Integrity Pact)"},
  {"S.No.": "22", "Particular": "Annexure -14 (Details of Drugs and Licenses)"},
  {"S.No.": "23", "Particular": "Authorization letter nominating a responsible person of the Bidder to attend the meetings like pre-bid & negotiation meeting. "},
  {"S.No.": "24", "Particular": "The bidder has to submit benchmark Purchase Order (PO)/Rate Reference/Recent Purchase Order in support of quoted items at the time of bid submission."},
  {"S.No.": "25", "Particular": "Vendor registration certificate(optional) -Vendors may apply for registration on the CGMSC portal; however, registration is not mandatory for participation in this tender."},
  {"S.No.": "26", "Particular": "Bid & bidder information (Download and extract Zip file from Tender Attachment)"}
]

class ComprehensiveEvaluator:
    def __init__(self):
        self.deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4o")
        self._client = None

    def _get_client(self) -> AzureOpenAI:
        """
        Lazily create the Azure OpenAI client.

        This keeps Excel export working even when Azure credentials are not configured.
        """
        if self._client is not None:
            return self._client

        api_key = os.getenv("AZURE_OPENAI_API_KEY")
        endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
        api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-12-01-preview")

        if not api_key or not endpoint:
            raise RuntimeError(
                "Azure OpenAI credentials are not configured (AZURE_OPENAI_API_KEY / AZURE_OPENAI_ENDPOINT)."
            )

        self._client = AzureOpenAI(
            api_key=api_key,
            api_version=api_version,
            azure_endpoint=endpoint,
        )
        return self._client

    def evaluate_all_documents(self, combined_text: str) -> list:
        """
        Uses Azure OpenAI to evaluate the comprehensive 26-point checklist 
        against the combined text of ALL uploaded documents.
        """
        # Truncate text to fit context window safely (~60,000 chars should be fine for gpt-4o, but adjust as needed)
        truncated_text = combined_text[:100000]

        schema_json = json.dumps(CRITERIA, indent=2)

        system_prompt = f"""You are an expert Tender Technical Evaluator for the CGMSCL Drug & Medicine procurement.
You will be provided with the combined OCR text extracted from ALL documents submitted by a bidding firm.

Your task is to fill out a 26-point evaluation checklist. For each criteria in the checklist, you must determine:
1. "Submitted": What information was found in the text regarding this requirement? Extract specific details like Certificate Numbers, validity dates, page numbers (if determinable), items quoted, and addresses.
   - If the criteria relates to Finance (e.g., Turnover, GST), output "TO BE EVALUATED BY FINANCE".
   - If it is not found, explain that it was not submitted.
   - SPECIAL RULE for S.No. "1" (Proof of Tender fees and EMD paid): Use concise wording in this style:
     "Paid Online: <Yes/No>; Submitted EMD Amount: <amount if found>; UTR Number: <utr if found>; Address: <address if found>; Seal: <Present/Not found>"
   - SPECIAL RULE for S.No. "2" (Copy of EMD exemption certificate if applicable): In "Submitted", output either:
     (a) "Applicable; Subject: <subject>; Tender Reference Number: <reference>; Address: <address>; Seal: <Present/Not found>"
     OR
     (b) "Not applicable"
   - SPECIAL RULE for S.No. "3" (Manufacturing Drug Licence): In "Submitted" you MUST build a compact paragraph that, where available, includes ALL of the following, clearly described and preferably in this order:
     (a) Manufacturing Licence number and its validity/expiry date and page number.
     (b) Drug codes and corresponding page numbers where the product permissions are shown.
     (c) Full manufacturer name and complete address block taken from the licence.
     The style should be similar to: "Mfg. Lic No.- <LIC_NO>, valid up to <DATE> (Page <X>). Product permission submitted for Drug Code <CODE(S)> on Page(s) <Y>. Address: <FULL ADDRESS>."  Use the actual details found in the text.
2. "As Per Requirement": Answer "Yes" or "No" based on whether the submitted documents meet the criteria. If it's a finance criteria, answer "Pending".
3. "Remark": Any discrepancy, warning, or additional note (e.g., "Item Code not Highlighted", "Submitted Blank Annexure-3, Clarify It").

Return ONLY a valid JSON array matching the exact 26 criteria. Each object in the array MUST have the following keys:
- "S.No." (string)
- "Particular" (string, exact description)
- "Submitted" (string, extraction result)
- "As Per Requirement" (string, "Yes", "No", or "Pending")
- "Remark" (string, or "-" if none)

Criteria List:
{schema_json}

Do NOT include any markdown formatting, code blocks, or explanations inside the response. Return raw JSON array ONLY."""

        user_prompt = f"Combined OCR Text from all documents:\n{truncated_text}\n\nPlease generate the 26-point evaluation array."

        try:
            client = self._get_client()
            response = client.chat.completions.create(
                model=self.deployment,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.1,
                max_tokens=4000,
            )

            raw = response.choices[0].message.content.strip()
            # Clean markdown wrappers if present
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
                if raw.endswith("```"):
                    raw = raw[:-3]
                raw = raw.strip()

            evaluations = json.loads(raw)
            return evaluations

        except Exception as e:
            # Offline/credentials-missing fallback: return a structurally valid 26-row checklist
            # so the frontend "Final Summary" is never empty.
            msg = str(e)[:200]
            print(f"  [WARN] Comprehensive evaluation unavailable, using fallback: {msg}", flush=True)
            return [
                {
                    "S.No.": c.get("S.No.", ""),
                    "Particular": c.get("Particular", ""),
                    "Submitted": "Not evaluated (offline).",
                    "As Per Requirement": "Pending",
                    "Remark": "-",
                }
                for c in CRITERIA
            ]

    def export_to_excel(self, evaluations: list, firm_name: str, output_path: str):
        """
        Exports the evaluation JSON to an Excel file structured like the given template.
        """
        default_template = os.path.join(os.path.dirname(__file__), "Comprehensive_Evaluation_Pfizer.xlsx")
        templates_dir_default = os.path.join(os.path.dirname(__file__), "templates")
        template_path = os.getenv("EVALUATION_TEMPLATE_PATH") or default_template

        # If the user dropped a template into backend/templates without setting env, prefer it.
        if not os.path.exists(template_path) and os.path.isdir(templates_dir_default):
            try:
                candidates = [
                    os.path.join(templates_dir_default, f)
                    for f in os.listdir(templates_dir_default)
                    if f.lower().endswith(".xlsx") and not f.startswith("~$")
                ]
                if candidates:
                    template_path = candidates[0]
            except Exception:
                template_path = default_template

        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # Common layout and styling so export stays consistent with tender format.
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        title_font = Font(name="Times New Roman", size=12, bold=True)
        header_font = Font(name="Times New Roman", size=11, bold=True)
        body_font = Font(name="Times New Roman", size=10)
        header_fill = PatternFill(start_color="9DB2C8", end_color="9DB2C8", fill_type="solid")

        ws.merge_cells("A1:E1")
        ws.merge_cells("A2:E2")
        ws.merge_cells("A3:E3")
        ws["A1"].value = "Tender Ref. No.232/CGMSCL/Drug&Medicine/2025-26, 05/02/2026"
        ws["A2"].value = "Cover A Evaluation"
        ws["A3"].value = f"Firm Name- {firm_name}"

        for row_idx in (1, 2):
            ws[f"A{row_idx}"].font = title_font
            ws[f"A{row_idx}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"A{row_idx}"].border = thin_border

        ws["A3"].font = header_font
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A3"].border = thin_border

        header_row = 4
        headers = ["S.No.", "Particular", "Submitted", "As Per Requirement", "Remark"]
        for col_idx, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 58
        ws.column_dimensions["C"].width = 52
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 34
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 30

        start_row = 5
        for i, eval_row in enumerate(evaluations):
            r = start_row + i
            ws.cell(row=r, column=1, value=eval_row.get("S.No.", ""))
            ws.cell(row=r, column=2, value=eval_row.get("Particular", ""))
            ws.cell(row=r, column=3, value=eval_row.get("Submitted", ""))
            ws.cell(row=r, column=4, value=eval_row.get("As Per Requirement", ""))
            ws.cell(row=r, column=5, value=eval_row.get("Remark", ""))

            # Keep long text readable, similar to the provided tender sheet style.
            row_values = [
                str(eval_row.get("Particular", "") or ""),
                str(eval_row.get("Submitted", "") or ""),
                str(eval_row.get("Remark", "") or ""),
            ]
            longest_text = max((len(text) for text in row_values), default=0)
            line_count_hint = max(1, (longest_text // 60) + 1)
            ws.row_dimensions[r].height = min(140, max(24, 16 * line_count_hint))

            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.font = body_font
                cell.border = thin_border
                if c in (1, 4):
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Clear stale values in template rows while preserving visual formatting.
        last_written = start_row + max(len(evaluations), 0) - 1
        if last_written < ws.max_row:
            for r in range(last_written + 1, ws.max_row + 1):
                for c in range(1, 6):
                    if ws.cell(row=r, column=c).value is not None:
                        ws.cell(row=r, column=c).value = None

        # Keep sheet name within Excel's 31-char limit and sanitize forbidden chars.
        safe_title = (firm_name or "Evaluation").replace("\\", "-").replace("/", "-").replace("*", "-")
        safe_title = safe_title.replace("[", "(").replace("]", ")").replace(":", "-").replace("?", "")
        safe_title = safe_title[:31]
        ws.title = safe_title or "Evaluation"
        wb.save(output_path)
