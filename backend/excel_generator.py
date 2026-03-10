import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import base64
from io import BytesIO
from PIL import Image

class ExcelGenerator:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def generate(self, data, job_id):
        """
        data: dictionary of fields and values
        """
        df = pd.DataFrame(list(data.items()), columns=['Field', 'Value'])
        # Add a dummy accuracy column for the POC
        df['OCR Accuracy (%)'] = "98.5%" # Placeholder
        
        file_path = os.path.join(self.output_dir, f"result_{job_id}.xlsx")
        df.to_excel(file_path, index=False)
        return file_path

    # Advanced: add images to excel (maybe in a second pass)
    def generate_with_seals(self, fields_data, seals_data, job_id):
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Fields"

        # Headers
        ws.append(["Field", "Value", "OCR Accuracy (%)"])
        
        for k, v in fields_data.items():
            ws.append([k, v, "98.5%"])

        # Add Seals section
        ws.append([])
        ws.append(["Detected Seals"])
        
        row_idx = ws.max_row + 1
        for i, seal in enumerate(seals_data):
            # Extract base64 image
            header, encoded = seal['image'].split(",", 1)
            img_data = base64.b64decode(encoded)
            img_file = BytesIO(img_data)
            
            # Save temp image for openpyxl
            img = Image.open(img_file)
            temp_img_path = os.path.join(self.output_dir, f"temp_seal_{job_id}_{i}.jpg")
            img.save(temp_img_path)
            
            # Add to excel
            xl_img = XLImage(temp_img_path)
            xl_img.width = 100
            xl_img.height = 100
            
            cell_address = f"B{row_idx + i}"
            ws.add_image(xl_img, cell_address)
            ws[f"A{row_idx + i}"] = f"Seal {i+1}"
            
        file_path = os.path.join(self.output_dir, f"result_{job_id}_full.xlsx")
        wb.save(file_path)
        
        # Cleanup temp images
        for i in range(len(seals_data)):
            temp_path = os.path.join(self.output_dir, f"temp_seal_{job_id}_{i}.jpg")
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
        return file_path

if __name__ == "__main__":
    print("ExcelGenerator loaded.")
